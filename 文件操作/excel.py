

import openpyxl
'''
# 操作现存的Excel文档；
# 1、加载Excel文档；
file = r'e:\lc.xlsx'
wb = openpyxl.load_workbook(file)
print(type(wb))

# 2、选择其中的工作表；
sheet = wb.get_sheet_by_name('lc01')
print(type(sheet))
print(sheet.title)

# 3、在表中获取行和列；
tuple(sheet['B2':'D34'])
for rowJect in sheet['B2':'D34']:
    for cellObj in rowJect:
        print(cellObj.coordinate, cellObj.value)
    print("---END OF ROW---")

# 4、从表中取得单元格；
n1 = sheet['D2']
print(n1.value)
'''

'''
# 新建新的Excel文档，并写入数据；
# 新建Excel文档；
new_excel = r'e:\newexcel.xlsx'
# 1、创建一个工作薄；
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
# 2、向工作薄写入数据；
sheet['A1'] = 'Name'
sheet['B1'] = 'Age'
sheet['A2'] = 'Zwj'
sheet['B2'] = '37'

# 3、修改工作薄的名字；
wb.save(new_excel)
# 输出结果；
print(f'{new_excel} created successfully.')
'''

# 向表中追加值；
# 加载Excel文档；
file = r'e:\newexcel.xlsx'
excel_file = openpyxl.load_workbook(file)
# 修改后保存为新的excel文档；
file2 = r'e:\newexcel2.xlsx'
# 显示所有工作表；
all_sheet = excel_file.sheetnames
print(all_sheet)
# 显示活动的表；
excel_sheet = excel_file.active
print(excel_sheet)
# 选择要操作的工作表；
sheet = excel_file['aa']
# 显示指定的工作表名称；
title_sheet = sheet.title
print(title_sheet)

# 向目标表中插入数据；
sheet['A3'] = 'Wangwu'
sheet['B3'] = 19
sheet['A4'] = 'QiHong'
sheet['B4'] = 33

# 修改后的excel文件另存为另一个文件；
excel_file.save(file2)
